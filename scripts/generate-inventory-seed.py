import json
import re
import sys
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl


SPECIAL_RESPONSIBLES = {"", "BAJA", "STOCK", "S/N", "S.N.", "SN", "N/A", "NA"}
DIRTY_AREAS = {"", "NO NAME", "NONAME", "N/A", "NO_NAME"}
ASSET_SHEETS = [
    "PC",
    "Desktop",
    "Lap Top",
    "Lap top",
    "Monitor",
    "Accesorios",
    "Teclado y raton",
    "Teclado y Raton",
    "UPS",
    "Tablets",
    "Tablet_Exp",
    "Audio & Video",
    "Audio y Video",
    "Multimedia",
    "Pantallas",
]


def clean(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if not text:
        return None
    return re.sub(r"\s+", " ", text)


def title_case(value):
    text = clean(value) or ""
    return " ".join(part.capitalize() for part in text.split())


def normalize_user_name(value):
    return title_case(value)


def normalize_area(value, stats=None):
    text = clean(value)
    if text is None or text.upper() in DIRTY_AREAS or text.upper().replace(" ", "") in DIRTY_AREAS:
        if stats is not None:
            stats["no_name_cleaned"] += 1
        return "SIN ÁREA"
    return text.upper()


def normalize_serial(value):
    text = clean(value)
    if text is None:
        return None
    return re.sub(r"\s+", "", text.upper())


def normalize_asset_name(value):
    return title_case(value)


def normalize_key(value):
    text = clean(value) or ""
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def asset_unique_key(asset):
    serial = asset.get("numeroSerie")
    if serial:
        return f"serial:{clean(serial)}"
    parts = [
        asset.get("nombre"),
        asset.get("marca"),
        asset.get("modelo"),
        asset.get("areaNombre"),
    ]
    return "fallback:" + "|".join(normalize_key(part) for part in parts)


def consolidate_similar_areas(users, assets, stats):
    counts = {}
    for record in list(users.values()) + assets:
        area = record.get("areaNombre")
        if area:
            counts[area] = counts.get(area, 0) + 1

    areas = sorted(counts)
    replacements = {}
    for index, area in enumerate(areas):
        if area in replacements:
            continue
        for other in areas[index + 1:]:
            if other in replacements:
                continue
            ratio = SequenceMatcher(None, normalize_key(area), normalize_key(other)).ratio()
            if ratio > 0.85:
                keep, replace = (area, other) if counts[area] >= counts[other] else (other, area)
                replacements[replace] = keep
                stats["areas_consolidated"] += 1
                print(f"Área '{replace}' consolidada en '{keep}'")

    for record in list(users.values()) + assets:
        area = record.get("areaNombre")
        if area in replacements:
            record["areaNombre"] = replacements[area]


def is_missing_code(value):
    text = (clean(value) or "").upper().replace(" ", "")
    return text in {"", "S/N", "SN", "S.N.", "S.N", "N/A", "NA", "NULL", "NONE"}


def tipo_from(equipo, sheet):
    source = f"{equipo or ''} {sheet}".upper()
    if "LAP" in source or "THINKPAD" in source:
        return "laptop"
    if "ALL IN ONE" in source or "AIO" in source:
        return "all-in-one"
    if "DESKTOP" in source or source.startswith("PC"):
        return "computadora"
    if "MONITOR" in source:
        return "monitor"
    if "IMPRESORA" in source or "PRINTER" in source:
        return "impresora"
    if "TABLET" in source or "IPAD" in source:
        return "tablet"
    if "UPS" in source:
        return "ups"
    if "TECLADO" in source or "MOUSE" in source or "RATON" in source or "ACCESOR" in source:
        return "accesorio"
    return "otro"


def slug_email(name):
    base = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii").lower()
    base = re.sub(r"[^a-z0-9]+", ".", base).strip(".")
    return f"{base or 'usuario'}@camaf.local"


def header_map(row):
    headers = {}
    duplicate_counts = {}
    for index, value in enumerate(row):
        key = normalize_key(value)
        if not key:
            continue
        duplicate_counts[key] = duplicate_counts.get(key, 0) + 1
        headers[key if duplicate_counts[key] == 1 else f"{key}{duplicate_counts[key]}"] = index
    return headers


def get(row, headers, *keys):
    for key in keys:
        index = headers.get(key)
        if index is not None and index < len(row):
            value = clean(row[index])
            if value is not None:
                return value
    return None


def main():
    if len(sys.argv) not in (3, 4):
        print("Usage: generate-inventory-seed.py <input.xlsx> <output.json> [unidad]", file=sys.stderr)
        sys.exit(2)

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    unidad = sys.argv[3] if len(sys.argv) == 4 else "Camaleón"
    workbook = openpyxl.load_workbook(input_path, data_only=True)
    stats = {
        "users_inserted": 0,
        "users_updated": 0,
        "assets_inserted": 0,
        "assets_updated": 0,
        "areas_consolidated": 0,
        "no_name_cleaned": 0,
    }

    users = {}
    if "USUARIOS" in workbook.sheetnames:
        sheet = workbook["USUARIOS"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = normalize_user_name(row[0] if len(row) > 0 else None)
            if not name:
                continue
            key = normalize_key(name)
            stats["users_updated" if key in users else "users_inserted"] += 1
            users[key] = {
                "nombre": name,
                "email": slug_email(name),
                "areaNombre": normalize_area(row[2] if len(row) > 2 else None, stats),
                "puesto": clean(row[3] if len(row) > 3 else None),
                "employeeId": clean(row[1] if len(row) > 1 else None),
                "unidad": unidad,
            }

    assets = []
    asset_index = {}
    for sheet_name in ASSET_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        headers = header_map(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not any(clean(value) for value in row[:12]):
                continue

            responsable = get(row, headers, "usuario")
            equipo = get(row, headers, "equipo")
            marca = get(row, headers, "marca")
            modelo = get(row, headers, "modelo")
            serial = get(row, headers, "deserie")
            internal_id = get(row, headers, "activo")
            inventory_number = get(row, headers, "activorlh", "activo2")
            departamento = get(row, headers, "departamento")
            ubicacion = get(row, headers, "ubicacion")
            observacion = get(row, headers, "observacion")
            fecha = get(row, headers, "fechadecompra")
            consecutivo = get(row, headers, "consecutivo")

            if internal_id and normalize_key(internal_id) in {
                normalize_key("RELACION DE STOCK DE EQUIPOS DE COMPUTO EN GENERAL"),
                normalize_key("BAJAS DE EQUIPOS Y EQUIPOS PARA PROCESO DE BAJA"),
            }:
                continue

            if not equipo and not marca and not modelo and is_missing_code(serial) and is_missing_code(internal_id):
                continue

            serial = None if is_missing_code(serial) else normalize_serial(serial)
            internal_id = None if is_missing_code(internal_id) else internal_id
            inventory_number = None if is_missing_code(inventory_number) else inventory_number

            area_name = normalize_area(departamento or ubicacion, stats)
            responsible_key = normalize_key(responsable)
            responsible_is_special = responsible_key in {normalize_key(item) for item in SPECIAL_RESPONSIBLES}
            if responsable and not responsible_is_special and responsible_key not in users:
                normalized_responsable = normalize_user_name(responsable)
                users[responsible_key] = {
                    "nombre": normalized_responsable,
                    "email": slug_email(normalized_responsable),
                    "areaNombre": area_name,
                    "puesto": None,
                    "employeeId": None,
                    "unidad": unidad,
                }
                stats["users_inserted"] += 1

            notes_parts = [
                f"Departamento: {departamento}" if departamento else None,
                f"Ubicacion: {ubicacion}" if ubicacion else None,
                f"Observacion: {observacion}" if observacion else None,
                f"RAM: {get(row, headers, 'ram')}" if get(row, headers, "ram") else None,
                f"Procesador: {get(row, headers, 'procesador')}" if get(row, headers, "procesador") else None,
                f"Disco duro: {get(row, headers, 'discoduro')}" if get(row, headers, "discoduro") else None,
                f"S.O.: {get(row, headers, 'so')}" if get(row, headers, "so") else None,
                f"Version: {get(row, headers, 'version')}" if get(row, headers, "version") else None,
                f"Origen Excel: {sheet_name}" if sheet_name else None,
            ]

            status = "baja" if normalize_key(responsable) == normalize_key("BAJA") or normalize_key(area_name) == normalize_key("BAJA") else "asignado"
            if not responsable or normalize_key(responsable) in {normalize_key("STOCK"), normalize_key("S/N"), normalize_key("S.N.")}:
                status = "activo"

            nombre = normalize_asset_name(" ".join(part for part in [equipo, marca, modelo] if part) or f"Activo {internal_id or serial or consecutivo}")
            asset_record = {
                "internalId": internal_id,
                "inventoryNumber": inventory_number,
                "nombre": nombre,
                "tipo": tipo_from(equipo, sheet_name),
                "marca": clean(marca),
                "modelo": clean(modelo),
                "numeroSerie": serial,
                "status": status,
                "areaNombre": area_name,
                "responsableNombre": None if responsible_is_special else normalize_user_name(responsable),
                "asignadoA": None if responsible_is_special else normalize_user_name(responsable),
                "unidad": unidad,
                "fechaAdquisicion": fecha,
                "notas": " | ".join(part for part in notes_parts if part) or None,
            }
            unique_key = asset_unique_key(asset_record)
            if unique_key in asset_index:
                assets[asset_index[unique_key]].update(asset_record)
                stats["assets_updated"] += 1
            else:
                asset_index[unique_key] = len(assets)
                assets.append(asset_record)
                stats["assets_inserted"] += 1

    used_internal_ids = set()
    for asset in assets:
        internal_id = asset.get("internalId")
        if not internal_id:
            continue
        if internal_id in used_internal_ids:
            asset["internalId"] = None
        else:
            used_internal_ids.add(internal_id)

    consolidate_similar_areas(users, assets, stats)

    output = {
        "source": input_path.name,
        "unidad": unidad,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "users": sorted(users.values(), key=lambda user: user["nombre"]),
        "assets": assets,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Usuarios insertados:   {stats['users_inserted']}")
    print(f"Usuarios actualizados: {stats['users_updated']}")
    print(f"Activos insertados:    {stats['assets_inserted']}")
    print(f"Activos actualizados:  {stats['assets_updated']}")
    print(f"Áreas consolidadas:    {stats['areas_consolidated']}")
    print(f"Registros \"NO NAME\" limpiados: {stats['no_name_cleaned']}")


if __name__ == "__main__":
    main()
